"""
EN: Excel reader service for QC.
VI: Dịch vụ đọc file Excel phục vụ kiểm tra chất lượng (QC).
"""

import os
import time
import logging
import openpyxl
import tempfile
import shutil
import re
from openpyxl.utils.exceptions import InvalidFileException
from typing import Dict, Any, Tuple, List
from src.models.config_model import AppConfig
from src.utils.formula_parser import ExcelFormulaEvaluator
from src.utils.core_utils import retry_io
from src.config.constants import IMPORT_NAME_KEYWORDS, IMPORT_FORMULA_KEYWORDS

logger = logging.getLogger(__name__)


@retry_io(retries=3, delay=1.0)
def _safe_load_workbook(file_path: str, data_only: bool = True):
    """EN: Load workbook with retry mechanism. VI: Mở tệp Excel kèm cơ chế thử lại."""
    return openpyxl.load_workbook(file_path, data_only=data_only)


def _check_fingerprint(ws, rule_str: str) -> bool:
    """EN: Check if the worksheet matches the fingerprint rule. VI: Kiểm tra xem sheet có khớp dấu vân tay không."""
    if not rule_str:
        return True

    conditions = [c.strip() for c in rule_str.split("&") if c.strip()]
    if not conditions:
        return True

    for condition in conditions:
        if "=" not in condition:
            continue
        parts = condition.split("=", 1)
        cell_coord = parts[0].strip().upper()
        expected_val = parts[1].strip().lower()
        try:
            actual_val = (
                str(ws[cell_coord].value).strip().lower()
                if ws[cell_coord].value is not None
                else ""
            )
            if expected_val not in actual_val:
                return False
        except (ValueError, TypeError, AttributeError, KeyError):
            return False

    return True


def get_sheet_names(file_path: str) -> List[str]:
    """EN: Get all sheet names from an Excel file safely. VI: Lấy danh sách tên sheet."""
    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    try:
        shutil.copy2(file_path, tmp_path)
        wb = _safe_load_workbook(tmp_path, data_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass


def get_sheet_preview(
    file_path: str, sheet_name: str, max_rows: int = 20
) -> List[List[Any]]:
    """EN: Get preview data from sheet. VI: Lấy tối đa 20 dòng để xem trước."""
    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    preview = []
    try:
        shutil.copy2(file_path, tmp_path)
        wb = _safe_load_workbook(tmp_path, data_only=True)
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                if r_idx >= max_rows:
                    break
                preview.append([str(c) if c is not None else "" for c in row])
        wb.close()
        return preview
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass


def extract_formulas_from_sheet(
    file_path: str,
    sheet_name: str,
    manual_name_col: int = -1,
    manual_form_col: int = -1,
    manual_start_row: int = -1,
) -> List[Tuple[str, str]]:
    """EN: Extract raw (name, formula) pairs from a specific sheet. VI: Trích xuất tên và công thức."""
    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    results = []
    try:
        shutil.copy2(file_path, tmp_path)
        wb = _safe_load_workbook(tmp_path, data_only=False)
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"Sheet '{sheet_name}' không tồn tại.")

        sheet = wb[sheet_name]
        name_col_idx = -1
        form_col_idx = -1
        start_row = 1

        if manual_name_col != -1 and manual_form_col != -1 and manual_start_row != -1:
            name_col_idx = manual_name_col
            form_col_idx = manual_form_col
            start_row = manual_start_row
        else:
            for r_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                f_name = -1
                f_form = -1
                for c_idx, cell_val in enumerate(row):
                    if cell_val:
                        val_str = str(cell_val).lower().strip()
                        if (
                            any(k in val_str for k in IMPORT_NAME_KEYWORDS)
                            and f_name == -1
                        ):
                            f_name = c_idx
                        elif (
                            any(k in val_str for k in IMPORT_FORMULA_KEYWORDS)
                            and f_form == -1
                        ):
                            f_form = c_idx
                if f_name != -1 and f_form != -1:
                    name_col_idx = f_name
                    form_col_idx = f_form
                    start_row = r_idx + 1
                    break

        if name_col_idx == -1 or form_col_idx == -1:
            wb.close()
            raise ValueError(
                "Không thể tự động nhận diện cột 'Tên' và 'Công thức' trong sheet này.\nVui lòng đảm bảo file có dòng tiêu đề chứa các từ khóa phù hợp."
            )

        seen_names = set()

        for row in sheet.iter_rows(min_row=start_row, values_only=True):
            if len(row) <= max(name_col_idx, form_col_idx):
                continue
            raw_name = row[name_col_idx]
            raw_form = row[form_col_idx]

            if not raw_name:
                continue

            clean_name = str(raw_name).strip()
            norm_name = normalize_name(clean_name)

            if norm_name in seen_names:
                wb.close()
                raise ValueError(
                    f"Lỗi: File của khách hàng chứa các cột có tên trùng lặp ('{clean_name}'). Vui lòng yêu cầu khách hàng sửa lại file trước khi nạp."
                )
            seen_names.add(norm_name)

            clean_form = sanitize_formula(raw_form)
            results.append((clean_name, clean_form))

        wb.close()
        return results
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass


def normalize_name(text: str) -> str:
    """EN: Normalize string by removing extra spaces and lowercasing. VI: Chuẩn hóa chuỗi."""
    return re.sub(r"\s+", " ", str(text)).strip().lower() if text else ""


def sanitize_formula(formula: Any) -> str:
    """EN: Clean external links from formula. VI: Xóa liên kết ngoại lai khỏi công thức."""
    if not formula:
        return ""
    form = str(formula).strip()
    if form.startswith("="):
        form = form[1:]
    form = re.sub(r"'?\[[^\]]+\][^'!]+'?!", "", form)
    return form.strip()


def process_import_mappings(
    extracted_data: List[Tuple[str, str]], target_names: List[str]
) -> Tuple[List[Tuple[str, str]], List[Tuple[str, str]], int]:
    """EN: Process extracted data into exact matches and pending aliases. VI: Phân loại dữ liệu trích xuất."""
    exact_matches = []
    pending_aliases = []
    skipped_count = 0

    for clean_name, clean_form in extracted_data:
        if not clean_form:
            skipped_count += 1
            continue

        norm_clean_name = normalize_name(clean_name)
        exact_match_name = None
        for t_name in target_names:
            if normalize_name(t_name) == norm_clean_name:
                exact_match_name = t_name
                break

        if exact_match_name:
            exact_matches.append((exact_match_name, clean_form))
        else:
            pending_aliases.append((clean_name, clean_form))

    return exact_matches, pending_aliases, skipped_count


def scan_file_qc(
    file_path: str, config: AppConfig, target_sheets: List[str] = None
) -> Tuple[bool, List[str], List[Dict[str, Any]]]:
    """
    EN: Scan input file, validate against config, and extract mapped values.
    VI: Quét file đầu vào, kiểm tra hợp lệ theo cấu hình, và trích xuất dữ liệu.
    Returns: (is_valid, list_of_errors, list_of_extracted_data_dicts)
    """
    errors = []
    extracted_data_list = []

    if not os.path.exists(file_path):
        errors.append(f"File không tồn tại: {file_path}")
        return False, errors, []

    wb = None
    try:
        start_time = time.time()
        logger.info(f"[Milestone] Bắt đầu quét file: {file_path}")

        wb = _safe_load_workbook(file_path, data_only=True)

        input_sheet_name = config.data.get("input_file", {}).get("sheet_name")
        sheets_to_scan = []

        if target_sheets is not None:
            for sn in target_sheets:
                if sn in wb.sheetnames:
                    sheets_to_scan.append(wb[sn])
                else:
                    errors.append(f"Không tìm thấy sheet '{sn}' trong tệp.")
        else:
            if not input_sheet_name:
                sheets_to_scan = [
                    ws for ws in wb.worksheets if ws.sheet_state == "visible"
                ]
            else:
                sheet_names = [s.strip() for s in input_sheet_name.split(",")]
                for sn in sheet_names:
                    if sn in wb.sheetnames:
                        sheets_to_scan.append(wb[sn])
                    else:
                        errors.append(f"Không tìm thấy sheet '{sn}' trong tệp.")

        if not sheets_to_scan and not errors:
            errors.append("Không có sheet nào hợp lệ được tìm thấy để quét.")

        for ws in sheets_to_scan:  # Lặp qua từng sheet để quét
            sheet_errors = []

            # --- Bắt đầu Logic Nhận diện Biểu mẫu (Fingerprint) ---
            active_config = config
            visited_profiles = set()
            match_found = False

            while active_config:
                profile_name = active_config.profile_name
                if profile_name in visited_profiles:
                    break  # Chống lặp vô hạn nếu người dùng cấu hình dự phòng vòng tròn
                visited_profiles.add(profile_name)

                fingerprint = active_config.data.get("fingerprint", "")
                if _check_fingerprint(ws, fingerprint):
                    match_found = True
                    break

                fallback_name = active_config.data.get("fallback_profile", "")
                if fallback_name and fallback_name != profile_name:
                    active_config = AppConfig(fallback_name)
                else:
                    break

            if not match_found:
                errors.append(
                    f"Sheet [{ws.title}] không khớp dấu vân tay của hồ sơ '{config.profile_name}' và các hồ sơ dự phòng."
                )
                continue
            # --- Kết thúc Logic Nhận diện ---

            evaluator = ExcelFormulaEvaluator(ws)
            extracted_data = {}

            for mapping in active_config.data.get("mappings", []):
                target_letter = mapping.get("target_col_letter")
                source_mapping = mapping.get("source_mapping")

                if not target_letter or not source_mapping:
                    continue

                try:
                    value = evaluator.evaluate(source_mapping)
                    if evaluator.errors:
                        sheet_errors.extend(evaluator.errors)
                    extracted_data[target_letter] = value
                except (
                    SyntaxError,
                    TypeError,
                    ValueError,
                    NameError,
                ) as e:
                    sheet_errors.append(
                        f"Sheet [{ws.title}] - Công thức '{source_mapping}': {e}"
                    )

            if sheet_errors:
                errors.extend(sheet_errors)
            else:
                extracted_data_list.append(extracted_data)

        end_time = time.time()
        logger.info(
            f"[Milestone] Hoàn tất trích xuất '{file_path}' trong {end_time - start_time:.2f} giây."
        )

    except PermissionError:
        errors.append(
            f"Tệp đang bị khóa (có thể đang mở trên Excel): {os.path.basename(file_path)}"
        )
    except InvalidFileException as e:
        errors.append(f"Định dạng tệp không hợp lệ: {e}")
    except (OSError, FileNotFoundError) as e:
        errors.append(f"Không thể đọc file Excel '{os.path.basename(file_path)}': {e}")
    finally:
        if wb is not None:
            wb.close()

    # Chỉ trả về False nếu không có mẻ nào được trích xuất thành công
    if not extracted_data_list:
        return (
            False,
            (
                errors
                if errors
                else ["Không trích xuất được dữ liệu từ bất kỳ sheet nào."]
            ),
            [],
        )

    return True, errors, extracted_data_list
