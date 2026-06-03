"""
EN: Extended unit tests for the Excel reader service to improve coverage.
VI: Kiem thu bo sung cho dich vu doc file Excel.
"""

import os
import pytest
import openpyxl
from unittest.mock import MagicMock, patch
from src.services.excel_reader_service import (
    _check_fingerprint,
    get_sheet_names,
    get_sheet_preview,
    scan_file_qc,
    sanitize_formula,
    normalize_name,
)
from src.models.config_model import AppConfig


def _create_xlsx(path, sheet_name="Sheet1", data=None):
    """Helper to create a simple xlsx file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    if data:
        for row_idx, row in enumerate(data, 1):
            for col_idx, val in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)
    wb.save(path)
    wb.close()


class TestCheckFingerprint:
    def test_empty_rule_returns_true(self):
        ws = MagicMock()
        assert _check_fingerprint(ws, "") is True

    def test_blank_conditions_returns_true(self):
        ws = MagicMock()
        assert _check_fingerprint(ws, "  &  ") is True

    def test_matching_fingerprint(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "BrewCard"
        assert _check_fingerprint(ws, "A1=brewcard") is True
        wb.close()

    def test_non_matching_fingerprint(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "OtherForm"
        assert _check_fingerprint(ws, "A1=brewcard") is False
        wb.close()

    def test_multiple_conditions_all_match(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "BrewCard"
        ws["B1"] = "v2"
        assert _check_fingerprint(ws, "A1=brewcard & B1=v2") is True
        wb.close()

    def test_multiple_conditions_one_fails(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "BrewCard"
        ws["B1"] = "v1"
        assert _check_fingerprint(ws, "A1=brewcard & B1=v2") is False
        wb.close()

    def test_cell_is_none(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        assert _check_fingerprint(ws, "A1=test") is False
        wb.close()

    def test_condition_without_equals_skipped(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "hello"
        assert _check_fingerprint(ws, "no_equals & A1=hello") is True
        wb.close()


class TestGetSheetNames:
    def test_returns_sheet_names(self, tmp_path):
        path = str(tmp_path / "test.xlsx")
        wb = openpyxl.Workbook()
        wb.create_sheet("Sheet2")
        wb.save(path)
        wb.close()

        names = get_sheet_names(path)
        assert "Sheet" in names
        assert "Sheet2" in names


class TestGetSheetPreview:
    def test_returns_preview_data(self, tmp_path):
        path = str(tmp_path / "preview.xlsx")
        _create_xlsx(path, "Data", [["A", "B"], [1, 2], [3, 4]])

        preview = get_sheet_preview(path, "Data", max_rows=2)
        assert len(preview) == 2
        assert preview[0] == ["A", "B"]
        assert preview[1] == ["1", "2"]

    def test_missing_sheet_returns_empty(self, tmp_path):
        path = str(tmp_path / "preview.xlsx")
        _create_xlsx(path, "Data")

        preview = get_sheet_preview(path, "NonExistent")
        assert preview == []


class TestScanFileQC:
    @pytest.fixture
    def mock_config(self, tmp_path):
        config = AppConfig("TestProfile")
        config.data["target_file"] = {
            "path": str(tmp_path / "master.xlsx"),
            "sheet_name": "Master",
        }
        config.data["input_file"] = {"sheet_name": "InputSheet"}
        config.data["header_row"] = "1"
        config.data["fingerprint"] = ""
        config.data["mappings"] = [
            {
                "target_col": "Batch",
                "target_col_letter": "A",
                "source_mapping": "B2",
            },
            {
                "target_col": "Value",
                "target_col_letter": "B",
                "source_mapping": "C2",
            },
        ]
        return config

    def test_permission_error(self, tmp_path, mock_config):
        """EN: Returns error when file is locked."""
        path = str(tmp_path / "locked.xlsx")
        _create_xlsx(path, "InputSheet")
        with patch(
            "src.services.excel_reader_service._safe_load_workbook",
            side_effect=PermissionError("file locked"),
        ):
            is_valid, errors, data = scan_file_qc(path, mock_config)
            assert is_valid is False
            assert any("khóa" in e.lower() or "locked" in e.lower() for e in errors)

    def test_invalid_file_format(self, tmp_path, mock_config):
        """EN: Returns error for invalid Excel format."""
        path = str(tmp_path / "bad.xlsx")
        with open(path, "w") as f:
            f.write("not an excel file")
        from openpyxl.utils.exceptions import InvalidFileException

        with patch(
            "src.services.excel_reader_service._safe_load_workbook",
            side_effect=InvalidFileException("bad format"),
        ):
            is_valid, errors, data = scan_file_qc(path, mock_config)
            assert is_valid is False
            assert any(
                "không hợp lệ" in e.lower() or "invalid" in e.lower() for e in errors
            )

    def test_scan_with_target_sheets_not_found(self, tmp_path, mock_config):
        """EN: Returns error when target sheet is not found in file."""
        path = str(tmp_path / "test.xlsx")
        _create_xlsx(path, "OtherSheet")
        is_valid, errors, data = scan_file_qc(
            path, mock_config, target_sheets=["MissingSheet"]
        )
        assert is_valid is False
        assert any("MissingSheet" in e for e in errors)

    def test_scan_no_valid_sheets(self, tmp_path, mock_config):
        """EN: Returns error when no valid sheets are found."""
        mock_config.data["input_file"]["sheet_name"] = "NonExistentSheet"
        path = str(tmp_path / "no_match.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "OtherSheet"
        wb.save(path)
        wb.close()

        is_valid, errors, data = scan_file_qc(path, mock_config)
        assert is_valid is False

    def test_scan_with_target_sheets_success(self, tmp_path, mock_config):
        """EN: Successfully scan specific target sheets."""
        path = str(tmp_path / "multi.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "InputSheet"
        ws["B2"] = "BATCH_X"
        ws["C2"] = 50
        wb.save(path)
        wb.close()

        is_valid, errors, data = scan_file_qc(
            path, mock_config, target_sheets=["InputSheet"]
        )
        assert is_valid is True
        assert len(data) > 0

    def test_scan_multiple_comma_separated_sheets(self, tmp_path, mock_config):
        """EN: Scan multiple comma-separated sheet names."""
        path = str(tmp_path / "multi.xlsx")
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1["B2"] = "BATCH_1"
        ws1["C2"] = 10
        ws2 = wb.create_sheet("Sheet2")
        ws2["B2"] = "BATCH_2"
        ws2["C2"] = 20
        wb.save(path)
        wb.close()

        mock_config.data["input_file"]["sheet_name"] = "Sheet1,Sheet2"
        is_valid, errors, data = scan_file_qc(path, mock_config)
        assert is_valid is True
        assert len(data) == 2

    def test_scan_fingerprint_mismatch(self, tmp_path, mock_config):
        """EN: Returns error when fingerprint does not match."""
        path = str(tmp_path / "fp.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "InputSheet"
        ws["A1"] = "WrongForm"
        ws["B2"] = "BATCH"
        ws["C2"] = 100
        wb.save(path)
        wb.close()

        mock_config.data["fingerprint"] = "A1=BrewCard"
        is_valid, errors, data = scan_file_qc(path, mock_config)
        assert is_valid is False
        assert any("vân tay" in e.lower() or "fingerprint" in e.lower() for e in errors)

    def test_scan_formula_evaluation_error(self, tmp_path, mock_config):
        """EN: Handles formula evaluation errors in mappings."""
        path = str(tmp_path / "formula_err.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "InputSheet"
        ws["B2"] = "BATCH_ERR"
        ws["C2"] = 0
        wb.save(path)
        wb.close()

        mock_config.data["mappings"].append(
            {
                "target_col": "Bad",
                "target_col_letter": "D",
                "source_mapping": "INVALID_FUNC(C2)",
            }
        )
        is_valid, errors, data = scan_file_qc(path, mock_config)
        assert is_valid is False

    def test_scan_mapping_without_source(self, tmp_path, mock_config):
        """EN: Mapping without source_mapping is skipped."""
        path = str(tmp_path / "nosrc.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "InputSheet"
        ws["B2"] = "BATCH"
        ws["C2"] = 10
        wb.save(path)
        wb.close()

        mock_config.data["mappings"].append(
            {"target_col": "Empty", "target_col_letter": "D", "source_mapping": ""}
        )
        is_valid, errors, data = scan_file_qc(path, mock_config)
        assert is_valid is True
