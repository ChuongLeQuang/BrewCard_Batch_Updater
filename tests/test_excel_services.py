"""
EN: Unit tests for Excel services.
VI: Kiểm thử đơn vị cho các dịch vụ xử lý Excel.
"""

import os
import pytest
import openpyxl
from src.services.excel_reader_service import (
    scan_file_qc,
    normalize_name,
    process_import_mappings,
    sanitize_formula,
)
from src.services.excel_sync_service import BrewSyncEngine
from src.models.config_model import AppConfig
from src.models.excel_data_model import BrewRecord


@pytest.fixture
def mock_config(tmp_path):
    config = AppConfig("TestProfile")
    config.data["target_file"] = {
        "path": str(tmp_path / "master.xlsx"),
        "sheet_name": "Master",
    }
    config.data["input_file"] = {"sheet_name": "InputSheet"}
    config.data["header_row"] = "1"
    config.data["mappings"] = [
        {"target_col": "Batch", "target_col_letter": "A", "source_mapping": "B2"},
        {"target_col": "Value", "target_col_letter": "B", "source_mapping": "C2"},
        {
            "target_col": "Calc",
            "target_col_letter": "C",
            "source_mapping": "C2 * 2",
            "format_type": "📊 Phần trăm (0.00%)",
        },  # Test tích hợp AST Parser
    ]
    return config


def test_scan_file_not_found(mock_config):
    is_valid, errors, data_list = scan_file_qc("fake_file.xlsx", mock_config)
    assert is_valid is False
    assert "không tồn tại" in errors[0].lower()


def test_scan_file_wrong_sheet(tmp_path, mock_config):
    file_path = tmp_path / "wrong_sheet.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "WrongName"
    wb.save(file_path)

    is_valid, errors, data_list = scan_file_qc(str(file_path), mock_config)
    assert is_valid is False
    assert "không tìm thấy sheet" in errors[0].lower()


def test_scan_file_success(tmp_path, mock_config):
    file_path = tmp_path / "valid.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "InputSheet"
    ws["B2"] = "BATCH_001"
    ws["C2"] = 100
    wb.save(file_path)

    is_valid, errors, data_list = scan_file_qc(str(file_path), mock_config)
    assert is_valid is True
    assert len(errors) == 0
    data = data_list[0]
    assert data["A"] == "BATCH_001"
    assert data["B"] == 100
    assert data["C"] == 200  # Đã qua Formula Parser tính toán


def test_sync_engine_insert_update(tmp_path, mock_config):
    engine = BrewSyncEngine(mock_config)
    records = [
        BrewRecord(batch_number="B002", data={"A": "B002", "B": 200, "C": 400}),
        BrewRecord(batch_number="B001", data={"A": "B001", "B": 100, "C": 200}),
    ]

    success = engine.sync_records(records)
    assert success is True

    target_path = mock_config.data["target_file"]["path"]
    assert os.path.exists(target_path)

    wb = openpyxl.load_workbook(target_path)
    ws = wb["Master"]
    assert ws["A1"].value == "Batch"  # Tự động tạo Header nếu file trống
    assert ws["A2"].value == "B002"  # Được nạp trước, ghi ở dòng 2
    assert ws["A3"].value == "B001"  # Được nạp sau, ghi ở dòng 3
    assert ws["C2"].number_format == "0.00%"  # Kiểm tra format được áp dụng
    wb.close()


def test_normalize_name():
    """EN: Test string normalization. VI: Kiểm thử tính năng chuẩn hóa chuỗi."""
    assert normalize_name("  Test   Name  ") == "test name"
    assert normalize_name("Test \n Name") == "test name"
    assert normalize_name("TEST") == "test"
    assert normalize_name(None) == ""
    assert normalize_name(123.45) == "123.45"


def test_process_import_mappings():
    """
    EN: Test separating extracted data into exact matches and pending aliases.
    VI: Kiểm thử chức năng phân tách dữ liệu thành danh sách khớp chuẩn và danh sách chờ ghép nối.
    """
    extracted_data = [
        ("Time In", "A1"),  # Khớp chuẩn (viết hoa/thường khác nhau)
        (" Time \n Out ", "B1"),  # Khớp chuẩn (có dấu xuống dòng, khoảng trắng)
        ("Unknown Col", "C1"),  # Không khớp -> Chuyển vào Pending
        ("Empty Form", ""),  # Công thức trống -> Bỏ qua
    ]
    target_names = ["time in", "time out", "other col"]

    exact, pending, skipped = process_import_mappings(extracted_data, target_names)

    # 1 dòng trống công thức phải bị bỏ qua
    assert skipped == 1

    # 2 dòng khớp chuẩn phải được nhận diện và lấy lại tên đích gốc
    assert len(exact) == 2
    assert exact[0] == ("time in", "A1")
    assert exact[1] == ("time out", "B1")

    # 1 dòng chưa rõ tên phải được đẩy vào pending
    assert len(pending) == 1
    assert pending[0] == ("Unknown Col", "C1")


def test_sanitize_formula():
    """EN: Test cleaning formulas. VI: Kiểm thử làm sạch công thức."""
    assert sanitize_formula("=[Data.xlsx]Sheet1!A1+B1") == "A1+B1"
    assert sanitize_formula("'[Data 2025.xlsx]Sheet 1'!A1") == "A1"
    assert sanitize_formula("=A1+B1") == "A1+B1"
    assert sanitize_formula(None) == ""


def test_sync_engine_inplace_update_preserves_other_columns(tmp_path, mock_config):
    """
    EN: Test that in-place updates preserve unmapped columns.
    VI: Kiểm thử đồng bộ tại chỗ giữ nguyên các cột phụ không có trong mapping.
    """
    # Cấu hình khóa chính cho Mapping
    mock_config.data["mappings"][0]["is_key"] = True

    target_path = mock_config.data["target_file"]["path"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master"
    ws["A1"] = "Batch"
    ws["B1"] = "Value"
    ws["C1"] = "Calc"
    ws["D1"] = "Extra1"
    ws["E1"] = "Extra2"

    ws["A2"] = "B001"
    ws["B2"] = 10
    ws["C2"] = 20
    ws["D2"] = "Ghi chú B001"
    ws["E2"] = "Nháp B001"

    ws["A3"] = "B003"
    ws["B3"] = 30
    ws["C3"] = 60
    ws["D3"] = "Ghi chú B003"
    ws["E3"] = "Nháp B003"

    wb.save(target_path)
    wb.close()

    engine = BrewSyncEngine(mock_config)
    records = [
        BrewRecord(batch_number="B002", data={"A": "B002", "B": 200, "C": 400}),
    ]

    success = engine.sync_records(records)
    assert success is True

    wb = openpyxl.load_workbook(target_path)
    ws = wb["Master"]

    # Tìm dòng của từng mẻ để assert chính xác, độc lập với việc nó ghi ở dòng mấy
    row_b001 = None
    row_b002 = None
    row_b003 = None

    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val == "B001":
            row_b001 = r
        elif val == "B002":
            row_b002 = r
        elif val == "B003":
            row_b003 = r

    assert row_b001 is not None
    assert row_b002 is not None
    assert row_b003 is not None

    # Mẻ B001 giữ nguyên giá trị mapping và cột phụ
    assert ws.cell(row=row_b001, column=2).value == 10
    assert ws.cell(row=row_b001, column=4).value == "Ghi chú B001"
    assert ws.cell(row=row_b001, column=5).value == "Nháp B001"

    # Mẻ B003 giữ nguyên giá trị mapping và cột phụ
    assert ws.cell(row=row_b003, column=2).value == 30
    assert ws.cell(row=row_b003, column=4).value == "Ghi chú B003"
    assert ws.cell(row=row_b003, column=5).value == "Nháp B003"

    # Mẻ B002 mới cập nhật đúng dữ liệu mapping và có cột phụ trống
    assert ws.cell(row=row_b002, column=2).value == 200
    assert ws.cell(row=row_b002, column=4).value is None
    assert ws.cell(row=row_b002, column=5).value is None

    wb.close()


def test_sync_engine_auto_backup(tmp_path, mock_config):
    """
    EN: Test that sync engine automatically backs up target file before writing.
    VI: Kiểm thử cơ chế tự động sao lưu tệp đích trước khi ghi đè.
    """
    target_path = mock_config.data["target_file"]["path"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master"
    ws["A1"] = "Batch"
    ws["A2"] = "B001"
    wb.save(target_path)
    wb.close()

    engine = BrewSyncEngine(mock_config)
    records = [BrewRecord(batch_number="B002", data={"A": "B002", "B": 200, "C": 400})]
    success = engine.sync_records(records)
    assert success is True

    backup_dir = "data/backups"
    assert os.path.exists(backup_dir)
    backups = os.listdir(backup_dir)
    assert len(backups) > 0
    assert any("TestProfile" in b and b.endswith(".xlsx") for b in backups)

    # Dọn dẹp backup
    for b in backups:
        try:
            os.remove(os.path.join(backup_dir, b))
        except OSError:
            pass
