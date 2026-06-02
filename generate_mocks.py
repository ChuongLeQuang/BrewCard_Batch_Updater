"""
EN: Helper script to generate mock Excel files for E2E testing.
VI: Script hỗ trợ tạo các file Excel giả lập phục vụ kiểm thử E2E.
"""

import os
import openpyxl
from datetime import datetime


def generate_mocks():
    out_dir = "tests/test_data"
    os.makedirs(out_dir, exist_ok=True)

    # 1. TẠO FILE TỔNG (MASTER)
    master_path = os.path.join(out_dir, "mock_master.xlsx")
    wb_master = openpyxl.Workbook()
    ws_master = wb_master.active
    ws_master.title = "BrewSync"
    ws_master.append(
        ["Batch", "Grist/Water Ratio", "Mash-In Time", "Yield (%)", "Date", "Status"]
    )
    ws_master.append(
        ["B9999", 0.15, 60, 0.95, datetime(2024, 1, 1), "Old Record"]
    )  # Dòng dữ liệu cũ
    wb_master.save(master_path)
    print(f"✅ Đã tạo: {master_path}")

    # 2. TẠO FILE INPUT HOÀN HẢO (CLEAN INSERT - Dành cho Test Happy Path)
    clean_path = os.path.join(out_dir, "mock_input_clean.xlsx")
    wb_clean = openpyxl.Workbook()
    ws_clean = wb_clean.active
    ws_clean.title = "Sheet1"

    # Vân tay và Số lô
    ws_clean["D3"] = "Order: 123"
    ws_clean["G3"] = "Batch: 10295"
    ws_clean["I3"] = "Date: 2024-03-20"
    ws_clean["H3"] = "10295"

    # Số liệu tính toán chuẩn xác
    ws_clean["E10"] = 100.0
    ws_clean["L15"] = 20.0
    ws_clean["L9"] = 50.0
    ws_clean["K10"] = 10.0
    ws_clean["F36"] = datetime(2024, 3, 20, 8, 0)
    ws_clean["F40"] = datetime(2024, 3, 20, 9, 30)

    # Bơm dữ liệu cho các ô đóng vai trò mẫu số (để tránh lỗi #DIV/0!)
    ws_clean["E9"] = 100.0
    ws_clean["E58"] = 100.0
    ws_clean["E78"] = 100.0
    ws_clean["E65"] = 100.0
    ws_clean["E97"] = 100.0

    # Bơm dữ liệu cho các ô đóng vai trò tử số (để số liệu đẹp)
    ws_clean["L8"] = 50.0
    ws_clean["E57"] = 50.0
    ws_clean["E83"] = 50.0
    ws_clean["D82"] = 50.0
    ws_clean["E99"] = 50.0

    wb_clean.save(clean_path)
    print(f"✅ Đã tạo: {clean_path}")

    # 3. TẠO FILE INPUT DỊ THƯỜNG (DIRTY MATH - Gây lỗi phép toán)
    dirty_path = os.path.join(out_dir, "mock_input_dirty.xlsx")
    wb_dirty = openpyxl.Workbook()
    ws_dirty = wb_dirty.active
    ws_dirty.title = "Sheet1"

    ws_dirty["D3"] = "Order: 124"
    ws_dirty["G3"] = "Batch: 10296"
    ws_dirty["I3"] = "Date: 2024-03-21"
    ws_dirty["H3"] = "10296"  # Số lô mới
    ws_dirty["E10"] = 0.0  # Cố tình gán = 0 để gây lỗi #DIV/0!
    ws_dirty["L15"] = 20.0
    ws_dirty["L9"] = 50.0
    ws_dirty["F36"] = datetime(2024, 3, 21, 8, 0)
    ws_dirty["F40"] = "Not a time"  # Cố tình điền chữ để gây lỗi #VALUE! / TypeError
    wb_dirty.save(dirty_path)
    print(f"✅ Đã tạo: {dirty_path}")

    # 4. TẠO FILE INPUT UPDATE ĐÈ (Trùng Số Lô B9999 có sẵn trong Master)
    update_path = os.path.join(out_dir, "mock_input_clean_update.xlsx")
    wb_update = openpyxl.Workbook()
    ws_update = wb_update.active
    ws_update.title = "Sheet1"

    ws_update["D3"] = "Order: 999"
    ws_update["G3"] = "Batch: B9999"
    ws_update["I3"] = "Date: 2024-05-01"
    ws_update["H3"] = "B9999"  # Số lô trùng Master
    ws_update["E10"] = 200.0
    ws_update["L15"] = 40.0
    wb_update.save(update_path)
    print(f"✅ Đã tạo: {update_path}")

    # 5. TẠO FILE INPUT CÔNG THỨC PHỨC TẠP (Logic IF, Time, %, Concat)
    complex_path = os.path.join(out_dir, "mock_input_complex_formulas.xlsx")
    wb_complex = openpyxl.Workbook()
    ws_complex = wb_complex.active
    ws_complex.title = "Sheet1"

    ws_complex["D3"] = "Order: 200"
    ws_complex["G3"] = "Batch: 1002"
    ws_complex["I3"] = "Date: 2024-05-02"
    ws_complex["H3"] = "1002"

    # Giả lập các công thức Excel thực tế
    ws_complex["A1"] = 100
    ws_complex["B1"] = 50
    ws_complex["E10"] = "=IF(A1>B1, 20%, 10%)"  # Sẽ trả về 0.2
    ws_complex["L15"] = "=AVERAGE(A1, B1)"  # Sẽ trả về 75
    ws_complex["L9"] = '=CONCATENATE("LOT-", H3)'  # Sẽ trả về LOT-1002
    wb_complex.save(complex_path)
    print(f"✅ Đã tạo: {complex_path}")

    # 6. TẠO FILE INPUT SAI BIỂU MẪU (Lỗi vân tay)
    wrong_form_path = os.path.join(out_dir, "mock_input_dirty_wrong_form.xlsx")
    wb_wrong = openpyxl.Workbook()
    ws_wrong = wb_wrong.active
    ws_wrong.title = "Sheet1"

    ws_wrong["D3"] = "Wrong Order Format"
    ws_wrong["G3"] = "Missing Keyword"
    ws_wrong["I3"] = "Date: 2024-05-03"
    ws_wrong["H3"] = "1004"
    wb_wrong.save(wrong_form_path)
    print(f"✅ Đã tạo: {wrong_form_path}")


if __name__ == "__main__":
    generate_mocks()
