"""
EN: Extended unit tests for the Excel sync service to improve coverage.
VI: Kiem thu bo sung cho dich vu dong bo Excel.
"""

import os
import pytest
import openpyxl
from src.services.excel_sync_service import BrewSyncEngine
from src.models.config_model import AppConfig
from src.models.excel_data_model import BrewRecord


@pytest.fixture
def base_config(tmp_path):
    config = AppConfig("SyncTest")
    config.data["target_file"] = {
        "path": str(tmp_path / "master.xlsx"),
        "sheet_name": "Master",
    }
    config.data["input_file"] = {"sheet_name": "Input"}
    config.data["header_row"] = "1"
    config.data["mappings"] = [
        {"target_col": "Batch", "target_col_letter": "A", "source_mapping": "B2"},
        {"target_col": "Value", "target_col_letter": "B", "source_mapping": "C2"},
    ]
    return config


class TestFindBatchColumn:
    def test_finds_batch_column_by_name(self, base_config):
        engine = BrewSyncEngine(base_config)
        assert engine.batch_col_letter == "A"

    def test_fallback_to_first_mapping(self, tmp_path):
        config = AppConfig("NoMatch")
        config.data["target_file"] = {
            "path": str(tmp_path / "master.xlsx"),
            "sheet_name": "Master",
        }
        config.data["mappings"] = [
            {"target_col": "Name", "target_col_letter": "C", "source_mapping": "A1"},
        ]
        engine = BrewSyncEngine(config)
        assert engine.batch_col_letter == "C"

    def test_fallback_empty_mappings(self, tmp_path):
        config = AppConfig("Empty")
        config.data["target_file"] = {
            "path": str(tmp_path / "master.xlsx"),
            "sheet_name": "Master",
        }
        config.data["mappings"] = []
        engine = BrewSyncEngine(config)
        assert engine.batch_col_letter == "A"


class TestSyncRecords:
    def test_insert_new_records_to_empty_file(self, tmp_path, base_config):
        engine = BrewSyncEngine(base_config)
        records = [
            BrewRecord(batch_number="B001", data={"A": "B001", "B": 100}),
            BrewRecord(batch_number="B002", data={"A": "B002", "B": 200}),
        ]
        success = engine.sync_records(records)
        assert success is True

        target_path = base_config.data["target_file"]["path"]
        wb = openpyxl.load_workbook(target_path)
        ws = wb["Master"]
        assert ws["A1"].value == "Batch"
        assert ws["A2"].value == "B001"
        assert ws["A3"].value == "B002"
        wb.close()

    def test_update_existing_records(self, tmp_path, base_config):
        """EN: Existing records are updated, not duplicated."""
        engine = BrewSyncEngine(base_config)
        records1 = [
            BrewRecord(batch_number="B001", data={"A": "B001", "B": 100}),
        ]
        engine.sync_records(records1)

        records2 = [
            BrewRecord(batch_number="B001", data={"A": "B001", "B": 999}),
        ]
        engine2 = BrewSyncEngine(base_config)
        success = engine2.sync_records(records2)
        assert success is True

        target_path = base_config.data["target_file"]["path"]
        wb = openpyxl.load_workbook(target_path)
        ws = wb["Master"]
        assert ws["B2"].value == 999
        assert ws.max_row == 2
        wb.close()

    def test_records_are_sorted_by_batch(self, tmp_path, base_config):
        """EN: Records are always sorted by batch number."""
        engine = BrewSyncEngine(base_config)
        records = [
            BrewRecord(batch_number="B003", data={"A": "B003", "B": 300}),
            BrewRecord(batch_number="B001", data={"A": "B001", "B": 100}),
            BrewRecord(batch_number="B002", data={"A": "B002", "B": 200}),
        ]
        success = engine.sync_records(records)
        assert success is True

        target_path = base_config.data["target_file"]["path"]
        wb = openpyxl.load_workbook(target_path)
        ws = wb["Master"]
        assert ws["A2"].value == "B001"
        assert ws["A3"].value == "B002"
        assert ws["A4"].value == "B003"
        wb.close()

    def test_sync_creates_new_sheet_in_existing_workbook(self, tmp_path, base_config):
        """EN: Creates a new sheet if it doesn't exist in an existing workbook."""
        target_path = base_config.data["target_file"]["path"]
        wb = openpyxl.Workbook()
        wb.active.title = "OtherSheet"
        wb.save(target_path)
        wb.close()

        engine = BrewSyncEngine(base_config)
        records = [
            BrewRecord(batch_number="B001", data={"A": "B001", "B": 100}),
        ]
        success = engine.sync_records(records)
        assert success is True

        wb = openpyxl.load_workbook(target_path)
        assert "Master" in wb.sheetnames
        ws = wb["Master"]
        assert ws["A2"].value == "B001"
        wb.close()

    def test_sync_with_format_type(self, tmp_path):
        """EN: Number format is applied to cells based on mapping format_type."""
        config = AppConfig("Formatted")
        config.data["target_file"] = {
            "path": str(tmp_path / "formatted.xlsx"),
            "sheet_name": "Master",
        }
        config.data["mappings"] = [
            {"target_col": "Batch", "target_col_letter": "A", "source_mapping": "B2"},
            {
                "target_col": "Pct",
                "target_col_letter": "B",
                "source_mapping": "C2",
                "format_type": "📊 Phần trăm (0.00%)",
            },
        ]
        engine = BrewSyncEngine(config)
        records = [
            BrewRecord(batch_number="B001", data={"A": "B001", "B": 0.85}),
        ]
        success = engine.sync_records(records)
        assert success is True

        wb = openpyxl.load_workbook(config.data["target_file"]["path"])
        ws = wb["Master"]
        assert ws["B2"].number_format == "0.00%"
        wb.close()

    def test_sync_returns_false_on_exception(self, tmp_path, base_config):
        """EN: Returns False when an exception occurs during sync."""
        engine = BrewSyncEngine(base_config)
        engine.target_path = "/nonexistent/path/master.xlsx"
        records = [
            BrewRecord(batch_number="B001", data={"A": "B001", "B": 100}),
        ]
        success = engine.sync_records(records)
        assert success is False
