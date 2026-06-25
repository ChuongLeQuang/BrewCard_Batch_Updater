"""
EN: Data Mappings Table Widget.
VI: Widget quản lý bảng lưới ánh xạ dữ liệu.
"""

import re
import os
from PyQt6.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QComboBox,
    QMessageBox,
    QDialog,
    QFileDialog,
    QInputDialog,
    QGroupBox,
)
from PyQt6.QtCore import Qt
from openpyxl.utils import column_index_from_string
from src.views.widget_noscroll_combobox import NoScrollComboBox
from src.views.widget_excel_mockup import WidgetExcelMockup
from src.views.config_dialog_alias import AliasMappingDialog
from src.models.alias_memory import AliasMemory
from src.views.dialog_column_selector import ColumnSelectorDialog
from src.config.constants import FORMAT_OPTIONS
from src.services.excel_reader_service import (
    get_sheet_names,
    extract_formulas_from_sheet,
    process_import_mappings,
    normalize_name,
    get_sheet_preview,
)


class ConfigWidgetMappings(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.letter_to_name = {}
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)

        grp_mappings = QGroupBox("🗺️ 3. Bản đồ ánh xạ dữ liệu (Mappings)", self)
        grp_mappings.setStyleSheet(
            "QGroupBox { font-weight: bold; color: #28A745; border: 1px solid #28A745; border-radius: 6px; margin-top: 12px; background-color: rgba(40, 167, 69, 0.03); } "
            "QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 5px; }"
        )
        layout_mappings = QVBoxLayout(grp_mappings)
        layout_mappings.setContentsMargins(12, 18, 12, 12)

        self.tbl_mappings = QTableWidget(0, 6, self)
        self.tbl_mappings.setHorizontalHeaderLabels(
            [
                "Cột đích (Chữ cái)",
                "Tên Cột đích (Tự động)",
                "Nguồn lấy dữ liệu / Công thức Excel",
                "🔑 Khóa Chính",
                "Định dạng hiển thị",
                "Xóa",
            ]
        )

        header = self.tbl_mappings.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive)
        self.tbl_mappings.setColumnWidth(0, 150)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.Interactive)
        self.tbl_mappings.setColumnWidth(4, 180)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.Fixed)
        self.tbl_mappings.setColumnWidth(5, 40)
        layout_mappings.addWidget(self.tbl_mappings)

        h_btns = QHBoxLayout()
        btn_add = QPushButton("➕ Thêm dòng ánh xạ", self)
        btn_add.clicked.connect(
            lambda: self._add_mapping_row(
                "", "", "", False, "📝 Mặc định", scroll_to_bottom=True
            )
        )
        btn_pick = QPushButton("🎯 Mở Excel ảo chọn Ô", self)
        btn_pick.clicked.connect(self._open_excel_mockup)
        btn_import = QPushButton("🔄 Nạp Công thức từ File...", self)
        btn_import.setStyleSheet("color: #E67E22; font-weight: bold;")
        btn_import.clicked.connect(self._import_formulas_from_file)
        btn_sort = QPushButton("🧹 Sắp xếp A-Z", self)
        btn_sort.clicked.connect(self.sort_and_refresh_ui)
        h_btns.addWidget(btn_add)
        h_btns.addWidget(btn_pick)
        h_btns.addWidget(btn_import)
        h_btns.addWidget(btn_sort)
        layout_mappings.addLayout(h_btns)

        layout.addWidget(grp_mappings)

        self.tbl_mappings.cellChanged.connect(self._on_cell_changed)

    def _on_cell_changed(self, row: int, col: int):
        item = self.tbl_mappings.item(row, col)
        if not item:
            return

        self.tbl_mappings.blockSignals(True)
        if col == 2 and item.text().strip():
            formatted = re.sub(
                r"\s*([()*/+<>=^&,\-])\s*", r"\1", item.text().strip().upper()
            )
            item.setText(formatted)
        elif col == 3:
            if item.checkState() == Qt.CheckState.Checked:
                for r in range(self.tbl_mappings.rowCount()):
                    if r != row:
                        other_item = self.tbl_mappings.item(r, 3)
                        if other_item:
                            other_item.setCheckState(Qt.CheckState.Unchecked)
        self.tbl_mappings.blockSignals(False)

    def update_headers(self, letter_to_name: dict):
        self.letter_to_name = letter_to_name
        for row in range(self.tbl_mappings.rowCount()):
            cmb = self.tbl_mappings.cellWidget(row, 0)
            if isinstance(cmb, QComboBox):
                current_val = cmb.currentText().upper()
                cmb.blockSignals(True)
                cmb.clear()
                sorted_letters = sorted(
                    list(self.letter_to_name.keys()), key=lambda l: (len(l), l)
                )
                cmb.addItems(sorted_letters)
                if current_val and cmb.findText(current_val) == -1:
                    cmb.addItem(current_val)
                cmb.setCurrentText(current_val)
                cmb.blockSignals(False)
                if current_val in self.letter_to_name:
                    self._update_target_name(row, current_val)

    def _update_target_name(self, row: int, col_letter: str):
        item = self.tbl_mappings.item(row, 1)
        if item:
            item.setText(self.letter_to_name.get(col_letter.strip().upper(), ""))

    def _update_target_name_by_cmb(self, cmb: QComboBox, col_letter: str):
        for row in range(self.tbl_mappings.rowCount()):
            if self.tbl_mappings.cellWidget(row, 0) == cmb:
                self._update_target_name(row, col_letter)
                break

    def _add_mapping_row(
        self,
        target_col: str,
        target_letter: str,
        source_mapping: str,
        is_key: bool = False,
        format_type: str = "📝 Mặc định",
        scroll_to_bottom: bool = False,
    ):
        row = self.tbl_mappings.rowCount()
        self.tbl_mappings.insertRow(row)

        cmb = NoScrollComboBox(self)
        cmb.setEditable(True)
        sorted_letters = sorted(
            list(self.letter_to_name.keys()), key=lambda l: (len(l), l)
        )
        cmb.addItems(sorted_letters)
        if target_letter and cmb.findText(target_letter.upper()) == -1:
            cmb.addItem(target_letter.upper())
        cmb.setCurrentText(target_letter.upper())
        self.tbl_mappings.setCellWidget(row, 0, cmb)

        item_name = QTableWidgetItem(target_col)
        item_name.setFlags(item_name.flags() & ~Qt.ItemFlag.ItemIsEditable)
        self.tbl_mappings.setItem(row, 1, item_name)
        self.tbl_mappings.setItem(row, 2, QTableWidgetItem(source_mapping))

        item_key = QTableWidgetItem()
        item_key.setFlags(Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled)
        item_key.setCheckState(
            Qt.CheckState.Checked if is_key else Qt.CheckState.Unchecked
        )
        self.tbl_mappings.setItem(row, 3, item_key)

        format_cmb = NoScrollComboBox(self)
        format_cmb.addItems(list(FORMAT_OPTIONS.keys()))
        if format_type and format_cmb.findText(format_type) != -1:
            format_cmb.setCurrentText(format_type)
        else:
            format_cmb.setCurrentIndex(0)
        self.tbl_mappings.setCellWidget(row, 4, format_cmb)

        btn_delete = QPushButton("❌", self)
        btn_delete.setStyleSheet(
            "color: red; border: none; font-size: 14px; padding: 2px;"
        )
        btn_delete.clicked.connect(lambda _, b=btn_delete: self._delete_mapping_row(b))
        self.tbl_mappings.setCellWidget(row, 5, btn_delete)

        cmb.currentTextChanged.connect(
            lambda text, c=cmb: self._update_target_name_by_cmb(c, text)
        )
        if target_letter:
            if target_letter.upper() in self.letter_to_name:
                self._update_target_name(row, target_letter)
            elif not target_col:
                self._update_target_name(row, target_letter)
        if scroll_to_bottom:
            self.tbl_mappings.scrollToBottom()

    def _delete_mapping_row(self, btn: QPushButton):
        for row in range(self.tbl_mappings.rowCount()):
            if self.tbl_mappings.cellWidget(row, 5) == btn:
                self.tbl_mappings.removeRow(row)
                break

    def _open_excel_mockup(self):
        dialog = WidgetExcelMockup(self)
        if dialog.exec() == QDialog.DialogCode.Accepted and dialog.selected_cell:
            curr = self.tbl_mappings.currentItem()
            if curr and curr.column() == 2:
                curr.setText(dialog.selected_cell)
            else:
                QMessageBox.information(
                    self,
                    "Tọa độ",
                    f"Đã chọn ô {dialog.selected_cell}\n(Ghi chú: Hãy click chọn trước một ô ở cột 'Nguồn' để phần mềm điền tự động)",
                )

    def _import_formulas_from_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Chọn File Công Thức từ Khách hàng",
            "",
            "Excel Files (*.xlsx *.xlsm);;All Files (*)",
        )
        if not file_path:
            return

        try:
            sheetnames = get_sheet_names(file_path)
            if not sheetnames:
                return

            sheet_name, ok = QInputDialog.getItem(
                self,
                "Chọn Sheet",
                "Vui lòng chọn Sheet chứa công thức:",
                sheetnames,
                0,
                False,
            )

            if not ok or not sheet_name:
                return

            try:
                extracted_data = extract_formulas_from_sheet(file_path, sheet_name)
            except ValueError as ve:
                if "Không thể tự động nhận diện" in str(ve):
                    preview_data = get_sheet_preview(file_path, sheet_name)
                    dialog = ColumnSelectorDialog(preview_data, self)
                    if dialog.exec() == QDialog.DialogCode.Accepted:
                        n_col, f_col, s_row = dialog.get_selection()
                        extracted_data = extract_formulas_from_sheet(
                            file_path, sheet_name, n_col, f_col, s_row
                        )
                    else:
                        return
                else:
                    raise ve

            imported_count = 0
            merged_count = 0

            self.tbl_mappings.setUpdatesEnabled(False)
            target_names_in_dest = list(self.letter_to_name.values())

            target_info = {}
            for r in range(self.tbl_mappings.rowCount()):
                item_name = self.tbl_mappings.item(r, 1)
                item_form = self.tbl_mappings.item(r, 2)
                if item_name and item_name.text().strip():
                    t_name = item_name.text().strip()
                    t_form = item_form.text().strip() if item_form else ""
                    if t_name not in target_info:
                        target_info[t_name] = t_form
                        if t_name not in target_names_in_dest:
                            target_names_in_dest.append(t_name)

            for t_name in target_names_in_dest:
                if t_name not in target_info:
                    target_info[t_name] = ""

            matched_targets_in_import = set()

            def apply_mapping(target_name: str, formula: str):
                nonlocal merged_count, imported_count
                found_r = -1
                norm_target = normalize_name(target_name)
                matched_targets_in_import.add(norm_target)
                for r in range(self.tbl_mappings.rowCount()):
                    item = self.tbl_mappings.item(r, 1)
                    if item and normalize_name(item.text()) == norm_target:
                        found_r = r
                        break

                fmt = "📝 Mặc định"
                nl = norm_target
                fl = formula.lower()
                if "%" in nl or "%" in fl or "ratio" in nl or "evap" in nl:
                    fmt = "📊 Phần trăm (0.00%)"
                elif "°c" in nl or "temp" in nl:
                    fmt = "🧮 Số thập phân (1)"
                elif (
                    "hl" in nl
                    or "plato" in nl
                    or "kg" in nl
                    or "ebc" in nl
                    or "ph" in nl
                    or "index" in nl
                    or "bar" in nl
                ):
                    fmt = "🧮 Số thập phân (2)"
                elif "min" in nl:
                    fmt = "🔢 Số nguyên"
                elif (
                    "time in" in nl
                    or "time out" in nl
                    or "start of" in nl
                    or "end of" in nl
                ):
                    fmt = "⏰ Ngày & Giờ (dd/mm/yyyy hh:mm)"
                elif "day" in nl or "date" in nl:
                    fmt = "📅 Ngày (dd/mm/yyyy)"

                if found_r != -1:
                    self.tbl_mappings.item(found_r, 2).setText(formula)
                    cmb = self.tbl_mappings.cellWidget(found_r, 4)
                    if (
                        isinstance(cmb, QComboBox)
                        and cmb.currentText() == "📝 Mặc định"
                        and fmt != "📝 Mặc định"
                    ):
                        cmb.setCurrentText(fmt)
                    merged_count += 1
                else:
                    best_letter = ""
                    for k, v in self.letter_to_name.items():
                        if normalize_name(v) == norm_target:
                            best_letter = k
                            break
                    self._add_mapping_row(
                        target_name, best_letter, formula, False, fmt, True
                    )
                    imported_count += 1

            exact_matches, pending_aliases, skipped_count = process_import_mappings(
                extracted_data, target_names_in_dest
            )

            alias_memory = AliasMemory()
            pending_after_memory = []

            for c_name, c_form in pending_aliases:
                mem_target = alias_memory.get_target(c_name)
                if mem_target and mem_target in target_names_in_dest:
                    exact_matches.append((mem_target, c_form))
                else:
                    pending_after_memory.append((c_name, c_form))

            pending_aliases = pending_after_memory

            for t_name, f_form in exact_matches:
                apply_mapping(t_name, f_form)

            if pending_aliases:
                available_targets_info = {
                    t: f
                    for t, f in target_info.items()
                    if normalize_name(t) not in matched_targets_in_import
                }
                dialog = AliasMappingDialog(
                    pending_aliases, available_targets_info, self
                )
                if dialog.exec() == QDialog.DialogCode.Accepted:
                    results = dialog.get_results()
                    for c_name, c_form in pending_aliases:
                        if c_name in results:
                            final_name, t_name = results[c_name]
                            apply_mapping(t_name, c_form)
                            alias_memory.add_alias(c_name, t_name)
                            if final_name != c_name:
                                alias_memory.add_alias(final_name, t_name)
                        else:
                            skipped_count += 1
                else:
                    skipped_count += len(pending_aliases)

            self.tbl_mappings.setUpdatesEnabled(True)

            QMessageBox.information(
                self,
                "Hoàn tất Nạp File",
                f"Đã xử lý xong!\n\n"
                f"🔄 Cập nhật/Ghi đè: {merged_count} dòng\n"
                f"➕ Thêm mới: {imported_count} dòng\n"
                f"⏭️ Bỏ qua (Trống/Giữ nguyên cũ): {skipped_count} dòng",
            )

        except ValueError as ve:
            QMessageBox.warning(self, "Không tìm thấy Tiêu đề", str(ve))
        except Exception as e:
            QMessageBox.critical(
                self, "Lỗi đọc file", f"Đã xảy ra lỗi khi nạp file: {str(e)}"
            )
        finally:
            self.tbl_mappings.setUpdatesEnabled(True)

    def load_mappings(self, mappings: list):
        self.tbl_mappings.setUpdatesEnabled(False)
        self.tbl_mappings.setRowCount(0)

        # Tự động sắp xếp trước khi hiển thị
        sorted_mappings = self._sort_mappings_list(mappings)

        has_key = any(m.get("is_key", False) for m in sorted_mappings)
        for i, m in enumerate(sorted_mappings):
            is_k = m.get("is_key", False)
            if not has_key and i == 0:
                is_k = True
            self._add_mapping_row(
                m.get("target_col", ""),
                m.get("target_col_letter", ""),
                m.get("source_mapping", ""),
                is_k,
                m.get("format_type", "📝 Mặc định"),
            )
        self.tbl_mappings.setUpdatesEnabled(True)

    def get_mappings(self) -> list:
        mappings = []
        for row in range(self.tbl_mappings.rowCount()):
            cmb = self.tbl_mappings.cellWidget(row, 0)
            tcl = (
                cmb.currentText().strip().upper() if isinstance(cmb, QComboBox) else ""
            )
            tc_item = self.tbl_mappings.item(row, 1)
            src_item = self.tbl_mappings.item(row, 2)
            key_item = self.tbl_mappings.item(row, 3)

            is_k = key_item.checkState() == Qt.CheckState.Checked if key_item else False

            format_cmb = self.tbl_mappings.cellWidget(row, 4)
            format_type = (
                format_cmb.currentText()
                if isinstance(format_cmb, QComboBox)
                else "📝 Mặc định"
            )

            mappings.append(
                {
                    "target_col": tc_item.text().strip() if tc_item else "",
                    "target_col_letter": tcl,
                    "source_mapping": src_item.text().strip() if src_item else "",
                    "is_key": is_k,
                    "format_type": format_type,
                }
            )
        # Tự động sắp xếp trước khi trả về
        return self._sort_mappings_list(mappings)

    def _sort_mappings_list(self, mappings: list) -> list:
        """
        EN: Sort mappings list alphabetically by target_col_letter (A-Z, AA-ZZ).
        VI: Sắp xếp danh sách ánh xạ theo thứ tự chữ cái của cột đích.
        """

        def get_col_index(item):
            letter = item.get("target_col_letter", "").strip().upper()
            if not letter:
                return 999999  # Đẩy các dòng chưa có cột đích xuống cuối cùng
            try:
                return column_index_from_string(letter)
            except ValueError:
                return 999999

        return sorted(mappings, key=get_col_index)

    def sort_and_refresh_ui(self):
        """
        EN: Sort current mappings and refresh table widget.
        VI: Sắp xếp bảng ánh xạ hiện tại và vẽ lại giao diện.
        """
        current_mappings = self.get_mappings()
        self.load_mappings(current_mappings)
