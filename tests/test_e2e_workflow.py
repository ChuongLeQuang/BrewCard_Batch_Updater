"""
EN: E2E Sandbox Test Workflow.
VI: Kiểm thử tích hợp toàn trình (E2E) trong môi trường Hộp cát (Sandbox).
"""

import os
import shutil
import pytest
import openpyxl
from src.models.config_model import AppConfig
from src.services.excel_reader_service import scan_file_qc
from src.services.excel_sync_service import BrewSyncEngine
from src.models.excel_data_model import BrewRecord
from src.config.constants import FORMAT_OPTIONS
from src.services.excel_reader_service import (
    extract_formulas_from_sheet,
    process_import_mappings,
)


@pytest.fixture
def sandbox_env(tmp_path):
    """
    EN: Setup sandbox environment with user's real mock files.
    VI: Thiết lập môi trường hộp cát với các file mẫu thật của người dùng.
    """
    test_data_dir = os.path.join(os.path.dirname(__file__), "test_data")
    files = {
        "master": "mock_master.xlsx",
        "clean": "mock_input_clean.xlsx",
        "dirty": "mock_input_dirty.xlsx",
        "update": "mock_input_clean_update.xlsx",
        "complex": "mock_input_complex_formulas.xlsx",
        "wrong_form": "mock_input_dirty_wrong_form.xlsx",
    }

    paths = {}
    for key, filename in files.items():
        src = os.path.join(test_data_dir, filename)
        if not os.path.exists(src):
            pytest.skip(
                f"Thiếu file dữ liệu mẫu ({filename}). Vui lòng chạy script generate_mocks.py."
            )

        dst = tmp_path / filename
        shutil.copy2(src, dst)
        paths[key] = str(dst)

    return paths


def test_e2e_happy_path(sandbox_env):
    """
    EN: Test Happy Path (Clean Data).
    VI: Kiểm thử luồng dữ liệu sạch (Happy Path), không có lỗi.
    """
    sandbox_master = sandbox_env["master"]
    sandbox_clean = sandbox_env["clean"]

    config = AppConfig()
    config.data["target_file"] = {"path": str(sandbox_master), "sheet_name": "Master"}
    config.data["fingerprint"] = ""
    config.data["fallback_profile"] = ""
    config.data["header_row"] = "1"
    config.data["header_row"] = "1"
    config.data["mappings"] = [
        {
            "target_col": "Batch Number",
            "target_col_letter": "A",
            "source_mapping": "H3",
            "format_type": "General",
            "is_key": True,
        },
        {
            "target_col": "Data",
            "target_col_letter": "B",
            "source_mapping": "E10",
            "format_type": "General",
        },
    ]

    is_valid, errors, extracted_data_list = scan_file_qc(str(sandbox_clean), config)

    assert is_valid is True, f"Quét QC thất bại. Lỗi chi tiết: {errors}"
    assert (
        len(errors) == 0
    ), f"Phát hiện lỗi không mong muốn trong luồng Happy Path: {errors}"
    assert len(extracted_data_list) > 0, "Không trích xuất được mẻ nấu nào."

    records = []
    engine = BrewSyncEngine(config)
    batch_col_letter = engine.batch_col_letter

    for data in extracted_data_list:
        batch_number = str(data.get(batch_col_letter, "")).strip()
        if batch_number:
            records.append(BrewRecord(batch_number=batch_number, data=data))

    assert len(records) > 0, f"Không tìm thấy Số lô."

    success = engine.sync_records(records)
    assert success is True, "Đồng bộ (Sync) vào file Master ảo thất bại."

    wb = openpyxl.load_workbook(str(sandbox_master), data_only=True)
    sheet_name = config.data["target_file"]["sheet_name"]
    assert (
        sheet_name in wb.sheetnames
    ), f"Không tìm thấy sheet '{sheet_name}' trong file Master."

    ws = wb[sheet_name]
    assert ws.max_row >= 2, "File Master ảo không có dữ liệu sau khi đồng bộ."

    from openpyxl.utils import column_index_from_string

    batch_col_idx = column_index_from_string(batch_col_letter)

    col_letter_to_format = {
        m["target_col_letter"]: FORMAT_OPTIONS.get(
            m.get("format_type", "📝 Mặc định"), "General"
        )
        for m in config.data.get("mappings", [])
    }

    error_strings = {"#DIV/0!", "#VALUE!", "#NAME?", "#ERROR!"}
    synced_rows_map = {}

    for r_idx in range(2, ws.max_row + 1):
        b_cell = ws.cell(row=r_idx, column=batch_col_idx)
        if b_cell.value is not None:
            b_num = str(b_cell.value).strip()
            synced_rows_map[b_num] = r_idx

            for c_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                if isinstance(cell.value, str) and cell.value in error_strings:
                    pytest.fail(
                        f"Happy Path không được phép chứa lỗi: {cell.value} tại {cell.coordinate}"
                    )

    for record in records:
        b_num = record.batch_number
        assert b_num in synced_rows_map, f"Không tìm thấy mẻ {b_num} trong file Master."

        actual_row_idx = synced_rows_map[b_num]
        for col_letter, expected_value in record.data.items():
            c_idx = column_index_from_string(col_letter)
            actual_cell = ws.cell(row=actual_row_idx, column=c_idx)

            actual_val = "" if actual_cell.value is None else actual_cell.value
            exp_val = "" if expected_value is None else expected_value

            if isinstance(exp_val, float) and isinstance(actual_val, float):
                assert actual_val == pytest.approx(
                    exp_val
                ), f"Sai số liệu tại {col_letter} mẻ {b_num}"
            else:
                assert actual_val == exp_val, f"Sai số liệu tại {col_letter} mẻ {b_num}"

        expected_format = col_letter_to_format.get(col_letter, "General")
        assert (
            actual_cell.number_format == expected_format
        ), f"Sai định dạng tại {col_letter} mẻ {b_num}"

    wb.close()

    output_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "output")
    os.makedirs(output_dir, exist_ok=True)
    shutil.copy2(
        str(sandbox_master), os.path.join(output_dir, "E2E_Happy_Path_Result.xlsx")
    )


def test_e2e_dirty_path(sandbox_env):
    """
    EN: Test Dirty Path (Error Data).
    VI: Kiểm thử luồng dữ liệu lỗi (Dirty Path). Hệ thống PHẢI TỪ CHỐI ghi.
    """
    sandbox_master = sandbox_env["master"]
    sandbox_dirty = sandbox_env["dirty"]

    config = AppConfig()
    config.data["target_file"] = {"path": str(sandbox_master), "sheet_name": "Master"}
    config.data["fingerprint"] = ""
    config.data["fallback_profile"] = ""
    config.data["mappings"] = [
        {
            "target_col": "Batch Number",
            "target_col_letter": "A",
            "source_mapping": "H3",
            "is_key": True,
        },
        {
            "target_col": "Error Data",
            "target_col_letter": "B",
            "source_mapping": "A1/0",  # Cố tình tạo lỗi chia cho 0 để đánh rớt QC
        },
    ]

    is_valid, errors, extracted_data_list = scan_file_qc(str(sandbox_dirty), config)

    # KỲ VỌNG 1: Hệ thống phải phát hiện lỗi và ĐÁNH RỚT file này (is_valid = False)
    assert (
        is_valid is False
    ), "Hệ thống BẮT BUỘC phải đánh rớt (Invalid) file chứa lỗi tính toán."

    # KỲ VỌNG 2: Phải ghi nhận được lỗi chi tiết để báo cho người dùng
    assert (
        len(errors) > 0
    ), "Luồng Dirty Path phải phát hiện ra lỗi (ZeroDivisionError, ValueError, v.v.)"

    # KỲ VỌNG 3: Dữ liệu của mẻ này KHÔNG được phép lọt qua chốt kiểm duyệt
    assert (
        len(extracted_data_list) == 0
    ), "Tuyệt đối không được trích xuất dữ liệu của mẻ nấu bị lỗi."

    wb = openpyxl.load_workbook(str(sandbox_master), data_only=True)

    error_strings = {"#DIV/0!", "#VALUE!", "#NAME?", "#ERROR!"}

    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=2, values_only=True):
            for cell_val in row:
                if isinstance(cell_val, str) and cell_val in error_strings:
                    pytest.fail(
                        f"LỖI NGHIÊM TRỌNG: Dữ liệu rác đã lọt qua chốt chặn và ghi vào file Tổng (Sheet '{ws.title}')!"
                    )

    wb.close()

    output_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "output")
    os.makedirs(output_dir, exist_ok=True)
    shutil.copy2(
        str(sandbox_master), os.path.join(output_dir, "E2E_Dirty_Path_Result.xlsx")
    )


def test_e2e_import_formulas_from_excel(tmp_path):
    """
    EN: E2E test for extracting and processing formulas from a real Excel file.
    VI: Kiểm thử E2E cho việc trích xuất và xử lý công thức từ file Excel thật.
    """
    # 1. Tạo file Excel giả lập (Đóng vai trò là File Công thức khách gửi)
    file_path = tmp_path / "customer_formulas.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Formulas"

    # Dòng Tiêu đề (Có chứa từ khóa)
    ws["A1"] = "STT"
    ws["B1"] = "Tên Chỉ Tiêu (Name)"
    ws["C1"] = "Công thức Excel (Formula)"

    # Dữ liệu mô phỏng các tình huống thực tế
    ws.append([1, "Mash-In Time", "=IF(A1>0, 10, 0)"])  # Khớp chuẩn, có dấu =
    ws.append([2, "Grist/Water Ratio", "=L15/E10%"])  # Khớp chuẩn, có toán tử %
    ws.append([3, "Unknown Column", "A1+B1"])  # Không khớp chuẩn -> Pending
    ws.append([4, "Empty Formula", None])  # Bỏ trống công thức -> Skip

    wb.save(file_path)

    # 2. Chạy hàm tự động trích xuất
    extracted_data = extract_formulas_from_sheet(str(file_path), "Formulas")
    assert len(extracted_data) == 4  # Lấy cả 4 dòng vì đều có Tên

    # 3. Chạy hàm phân loại dữ liệu (Mapping)
    target_names = ["Mash-in time", "grist/water ratio", "Deep Cut"]
    exact, pending, skipped = process_import_mappings(extracted_data, target_names)

    # 4. Nghiệm thu kết quả phân loại
    assert skipped == 1  # Dòng 'Empty Formula' bị bỏ qua do không có công thức

    assert len(exact) == 2
    exact_dict = dict(exact)
    assert exact_dict["Mash-in time"] == "IF(A1>0, 10, 0)"  # Đã gọt dấu =
    assert exact_dict["grist/water ratio"] == "L15/E10%"  # Đã gọt dấu =, giữ nguyên %

    assert len(pending) == 1
    assert pending[0] == ("Unknown Column", "A1+B1")  # Đẩy vào Hộp thoại chờ


def test_e2e_update_overwrite(sandbox_env):
    """
    EN: Test updating/overwriting an existing batch record.
    VI: Kiểm thử luồng Ghi đè cập nhật mẻ nấu đã tồn tại.
    """
    sandbox_master = sandbox_env["master"]
    sandbox_update = sandbox_env["update"]

    config = AppConfig()
    config.data["target_file"] = {"path": str(sandbox_master), "sheet_name": "Master"}
    config.data["fingerprint"] = ""
    config.data["fallback_profile"] = ""
    config.data["header_row"] = "1"
    # Cố tình set mapping trỏ đúng vào Cột A (Số lô) để test
    config.data["mappings"] = [
        {
            "target_col": "Batch Number",
            "target_col_letter": "A",
            "source_mapping": "H3",
            "format_type": "General",
            "is_key": True,
        },
        {
            "target_col": "Grist/Water Ratio",
            "target_col_letter": "B",
            "source_mapping": "E10",
            "format_type": "General",
        },
    ]

    is_valid, errors, extracted = scan_file_qc(sandbox_update, config)
    assert is_valid is True, f"QC thất bại: {errors}"

    records = [BrewRecord(batch_number=str(d.get("A", "")), data=d) for d in extracted]
    engine = BrewSyncEngine(config)
    success = engine.sync_records(records)

    assert success is True

    wb = openpyxl.load_workbook(sandbox_master, data_only=True)
    ws = wb[config.data["target_file"]["sheet_name"]]

    # File Master ban đầu có 1 Header + 1 Dòng dữ liệu (Lô B9999).
    # File Update cũng chứa lô B9999. Do đó max_row vẫn chỉ được phép là 2 (Không chèn dòng mới).
    batch_count = sum(
        1 for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=1).value
    )
    assert (
        batch_count == 1
    ), "Lỗi nghiêm trọng: Hệ thống đẻ thêm dòng mới thay vì ghi đè dòng cũ!"

    # Kiểm tra xem dữ liệu có thực sự được cập nhật chưa
    assert ws["A2"].value == "B9999"
    assert ws["B2"].value == 200.0  # Giá trị E10 mới từ mock_input_clean_update
    wb.close()


def test_e2e_complex_formulas(sandbox_env):
    """
    EN: Test parsing and calculating complex AST formulas.
    VI: Kiểm thử việc phân dịch các công thức phức tạp qua AST.
    """
    sandbox_complex = sandbox_env["complex"]

    config = AppConfig()
    config.data["fingerprint"] = ""
    config.data["fallback_profile"] = ""
    config.data["header_row"] = "1"
    config.data["mappings"] = [
        {
            "target_col": "Batch",
            "target_col_letter": "A",
            "source_mapping": "H3",
            "format_type": "General",
        },
        {
            "target_col": "Result IF",
            "target_col_letter": "B",
            "source_mapping": "IF(A1>B1, 20%, 10%)",
            "format_type": "General",
        },
        {
            "target_col": "Result AVG",
            "target_col_letter": "C",
            "source_mapping": "AVERAGE(A1, B1)",
            "format_type": "General",
        },
        {
            "target_col": "Result CONCAT",
            "target_col_letter": "D",
            "source_mapping": 'CONCATENATE("LOT-", H3)',
            "format_type": "General",
        },
    ]

    is_valid, errors, extracted = scan_file_qc(sandbox_complex, config)
    assert is_valid is True, f"Lỗi dịch công thức AST: {errors}"
    assert len(extracted) == 1

    data = extracted[0]
    assert data["A"] == "1002"
    assert data["B"] == 0.2  # =IF(A1>B1, 20%, 10%) với A1=100, B1=50
    assert data["C"] == 75.0  # =AVERAGE(A1, B1) với A1=100, B1=50
    assert data["D"] == "LOT-1002"  # =CONCATENATE("LOT-", H3) với H3=1002


def test_e2e_wrong_fingerprint(sandbox_env):
    """
    EN: Test rejection of files that do not match the fingerprint.
    VI: Kiểm thử việc từ chối các file không khớp dấu vân tay nhận diện.
    """
    sandbox_wrong = sandbox_env["wrong_form"]

    config = AppConfig()
    config.data["fingerprint"] = "D3=Order: & G3=Batch:"
    config.data["fallback_profile"] = ""

    is_valid, errors, extracted = scan_file_qc(sandbox_wrong, config)

    assert is_valid is False
    assert len(extracted) == 0
    assert any(
        "không khớp" in err.lower()
        or "vân tay" in err.lower()
        or "biểu mẫu" in err.lower()
        for err in errors
    ), "File sai biểu mẫu nhưng không bị báo lỗi vân tay!"
