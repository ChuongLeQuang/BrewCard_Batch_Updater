"""
EN: Shared file utilities for safe Excel workbook handling.
VI: Tiện ích dùng chung để xử lý file Excel an toàn.
"""

import os
import tempfile
import shutil
from contextlib import contextmanager
from typing import List

import openpyxl

from src.utils.core_utils import retry_io


@retry_io(retries=3, delay=1.0)
def safe_load_workbook(file_path: str, read_only: bool = False, data_only: bool = True):
    """
    EN: Load an Excel workbook with automatic retry on I/O errors.
    VI: Mở tệp Excel kèm cơ chế thử lại tự động khi gặp lỗi I/O.
    """
    return openpyxl.load_workbook(file_path, read_only=read_only, data_only=data_only)


@contextmanager
def temp_workbook(file_path: str, read_only: bool = False, data_only: bool = True):
    """
    EN: Context manager that copies an Excel file to a temp location,
        opens it safely, and cleans up automatically.
    VI: Context manager sao chép file Excel sang thư mục tạm,
        mở an toàn, và tự dọn dẹp sau khi xong.

    Usage::

        with temp_workbook(path, data_only=True) as wb:
            ws = wb[sheet_name]
            ...
    """
    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    try:
        shutil.copy2(file_path, tmp_path)
        wb = safe_load_workbook(tmp_path, read_only=read_only, data_only=data_only)
        try:
            yield wb
        finally:
            wb.close()
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass


def find_batch_column(mappings: List[dict]) -> str:
    """
    EN: Find the column letter for the batch number from a list of mappings.
    VI: Tìm chữ cái cột Số lô từ danh sách ánh xạ.
    """
    for m in mappings:
        if "batch" in m.get("target_col", "").lower():
            return m.get("target_col_letter", "A")
    return mappings[0].get("target_col_letter", "A") if mappings else "A"
