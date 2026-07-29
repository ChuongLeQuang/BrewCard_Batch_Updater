"""
EN: Profile & Target File Configuration Widget.
VI: Widget quản lý cấu hình Profile và File Đích.
"""

import os
import re
import openpyxl
from openpyxl.utils import get_column_letter
from PyQt6.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QPushButton,
    QComboBox,
    QFileDialog,
    QMessageBox,
    QGroupBox,
)
from PyQt6.QtCore import Qt, pyqtSignal
from src.views.widget_noscroll_combobox import NoScrollComboBox
from src.utils.core_utils import retry_io


class ConfigWidgetProfile(QWidget):
    headers_refreshed = pyqtSignal(dict)  # Phát ra dict letter_to_name
    profile_changed = pyqtSignal(str)  # Phát ra tên profile mới
    delete_requested = pyqtSignal(str)  # Phát ra tên profile cần xóa

    def __init__(self, parent=None):
        super().__init__(parent)
        self.available_headers = {}
        self.letter_to_name = {}
        self._setup_ui()
        self._setup_connections()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)

        # Dòng 0: Quản lý Đa cấu hình (Profiles)
        h_profile = QHBoxLayout()
        h_profile.addWidget(QLabel("📂 Hồ sơ Cấu hình (Profile):", self))
        self.cmb_profile = NoScrollComboBox(self)
        self.cmb_profile.setEditable(True)
        h_profile.addWidget(self.cmb_profile, stretch=1)
        btn_delete_profile = QPushButton("🗑️ Xóa Profile", self)
        btn_delete_profile.setStyleSheet("color: #FF3333;")
        btn_delete_profile.clicked.connect(
            lambda: self.delete_requested.emit(self.cmb_profile.currentText().strip())
        )
        h_profile.addWidget(btn_delete_profile)
        layout.addLayout(h_profile)

        # Khung 1: Thiết lập File Tổng (Master File)
        grp_master = QGroupBox("📦 1. Thiết lập File Tổng (Nơi dồn dữ liệu vào)", self)
        grp_master.setStyleSheet(
            "QGroupBox { font-weight: bold; color: #2A82DA; border: 1px solid #2A82DA; border-radius: 6px; margin-top: 12px; background-color: rgba(42, 130, 218, 0.03); } "
            "QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 5px; }"
        )
        layout_master = QVBoxLayout(grp_master)
        layout_master.setContentsMargins(12, 18, 12, 12)

        h_target = QHBoxLayout()
        h_target.addWidget(QLabel("Tệp Excel Đích:", self))
        self.txt_target_path = QLineEdit(self)
        h_target.addWidget(self.txt_target_path)
        btn_browse = QPushButton("📁 Chọn tệp", self)
        btn_browse.clicked.connect(self._browse_target_file)
        h_target.addWidget(btn_browse)
        layout_master.addLayout(h_target)

        h_sheet = QHBoxLayout()
        h_sheet.addWidget(QLabel("Lưu vào Sheet:", self))
        self.cmb_sheet_name = NoScrollComboBox(self)
        self.cmb_sheet_name.setEditable(True)
        h_sheet.addWidget(self.cmb_sheet_name, stretch=1)
        h_sheet.addWidget(QLabel("Dòng tiêu đề (vd: 5):", self))
        self.txt_header_row = QLineEdit(self)
        self.txt_header_row.setFixedWidth(80)
        h_sheet.addWidget(self.txt_header_row)
        layout_master.addLayout(h_sheet)
        layout.addWidget(grp_master)

        # Khung 2: Tự động Nhận diện File Con (Input Files)
        grp_input = QGroupBox(
            "🤖 2. Tự động Nhận diện File Con (Kéo thả ở Tab QC)", self
        )
        grp_input.setStyleSheet(
            "QGroupBox { font-weight: bold; color: #E67E22; border: 1px solid #E67E22; border-radius: 6px; margin-top: 12px; background-color: rgba(230, 126, 34, 0.03); } "
            "QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 5px; }"
        )
        layout_input = QVBoxLayout(grp_input)
        layout_input.setContentsMargins(12, 18, 12, 12)

        h_in_sheet = QHBoxLayout()
        h_in_sheet.addWidget(QLabel("Tên Sheet lấy dữ liệu:", self))
        self.txt_input_sheet = QLineEdit(self)
        self.txt_input_sheet.setPlaceholderText(
            "Để trống = Tự động quét tất cả các sheet có trong file"
        )
        h_in_sheet.addWidget(self.txt_input_sheet)
        layout_input.addLayout(h_in_sheet)

        h_fp1 = QHBoxLayout()
        h_fp1.addWidget(
            QLabel(
                "🔍 <b>Dấu hiệu nhận diện</b> <i>(VD: G3=Batch & D3=Order)</i>:", self
            )
        )
        self.txt_fingerprint_rule = QLineEdit(self)
        self.txt_fingerprint_rule.setPlaceholderText("Để trống = Bỏ qua kiểm tra Form")
        h_fp1.addWidget(self.txt_fingerprint_rule, stretch=1)
        layout_input.addLayout(h_fp1)

        h_fp2 = QHBoxLayout()
        h_fp2.addWidget(QLabel("🔄 Nếu không khớp, tự động dùng Form dự phòng:", self))
        self.cmb_fallback_profile = NoScrollComboBox(self)
        h_fp2.addWidget(self.cmb_fallback_profile, stretch=1)
        layout_input.addLayout(h_fp2)

        self.lbl_chain_preview = QLabel("👉 Luồng: ...", self)
        self.lbl_chain_preview.setStyleSheet(
            "color: #28A745; font-style: italic; margin-left: 10px;"
        )
        layout_input.addWidget(self.lbl_chain_preview)
        layout.addWidget(grp_input)

    def _setup_connections(self):
        self.cmb_profile.currentIndexChanged.connect(self._on_profile_index_changed)
        self.cmb_sheet_name.currentTextChanged.connect(self._refresh_headers)
        self.txt_header_row.editingFinished.connect(self._refresh_headers)
        self.txt_target_path.editingFinished.connect(
            lambda: self._populate_sheet_names(self.txt_target_path.text())
        )

        # Cập nhật sơ đồ luồng khi đổi Combobox
        self.cmb_profile.currentTextChanged.connect(self._update_chain_preview)
        self.cmb_fallback_profile.currentTextChanged.connect(self._update_chain_preview)

    def _on_profile_index_changed(self, index: int):
        # Chỉ tải cấu hình khi chọn một profile hợp lệ từ danh sách (index >= 0).
        # Tránh tải lại làm mất dữ liệu giao diện khi người dùng đang gõ tên profile mới (index = -1).
        if index >= 0:
            profile_name = self.cmb_profile.itemText(index).strip()
            if profile_name:
                self.profile_changed.emit(profile_name)

    def _update_chain_preview(self):
        current_prof = self.cmb_profile.currentText().strip()
        fallback_prof = self.cmb_fallback_profile.currentText().strip()
        if not current_prof:
            self.lbl_chain_preview.setText("")
            return

        if fallback_prof and fallback_prof != current_prof:
            self.lbl_chain_preview.setText(
                f"👉 Luồng quét: Kiểm tra [{current_prof}] ➡️ (Nếu sai) ➡️ Chuyển sang [{fallback_prof}]"
            )
        else:
            self.lbl_chain_preview.setText(
                f"👉 Luồng quét: Kiểm tra [{current_prof}] ➡️ (Nếu sai) ➡️ ❌ Báo Lỗi Ngay"
            )

    @retry_io(retries=3, delay=1.0)
    def _safe_load_workbook(
        self, file_path: str, read_only: bool = True, data_only: bool = False
    ):
        return openpyxl.load_workbook(
            file_path, read_only=read_only, data_only=data_only
        )

    def _browse_target_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Chọn tệp Excel đích",
            "",
            "Excel Files (*.xlsx *.xlsm);;All Files (*)",
        )
        if file_path:
            self.txt_target_path.setText(file_path)
            self._populate_sheet_names(file_path)
            self._refresh_headers()

    def _populate_sheet_names(self, file_path: str, default_sheet: str = ""):
        self.cmb_sheet_name.blockSignals(True)
        self.cmb_sheet_name.clear()
        sheets = []
        if os.path.exists(file_path):
            if file_path.lower().endswith((".xlsx", ".xlsm")):
                wb = None
                try:
                    wb = self._safe_load_workbook(file_path)
                    sheets = wb.sheetnames
                except (
                    openpyxl.utils.exceptions.InvalidFileException,
                    FileNotFoundError,
                    PermissionError,
                ) as e:
                    QMessageBox.warning(
                        self,
                        "Lỗi đọc file",
                        f"Không thể đọc danh sách sheet.\nChi tiết lỗi: {str(e)}",
                    )
                finally:
                    if wb is not None:
                        wb.close()
            elif file_path.lower().endswith((".xlsb", ".xls")):
                QMessageBox.warning(
                    self,
                    "Định dạng không hỗ trợ",
                    "Thư viện chỉ hỗ trợ .xlsx hoặc .xlsm.",
                )

        if sheets:
            self.cmb_sheet_name.addItems(sheets)
            if default_sheet in sheets:
                self.cmb_sheet_name.setCurrentText(default_sheet)
        else:
            if default_sheet:
                self.cmb_sheet_name.addItem(default_sheet)
        self.cmb_sheet_name.blockSignals(False)

    def _refresh_headers(self):
        # Giữ lại toàn bộ logic trích xuất header cũ
        target_path = self.txt_target_path.text().strip()
        sheet_name = self.cmb_sheet_name.currentText().strip()
        try:
            h_text = self.txt_header_row.text().strip()
            if "-" in h_text:
                start_r, end_r = map(int, h_text.split("-"))
                if start_r > end_r:
                    start_r, end_r = end_r, start_r
            else:
                start_r = end_r = int(h_text)
        except ValueError:
            return

        self.available_headers.clear()
        self.letter_to_name.clear()

        if os.path.exists(target_path) and target_path.lower().endswith(
            (".xlsx", ".xlsm")
        ):
            wb = None
            try:
                wb = self._safe_load_workbook(
                    target_path, read_only=True, data_only=True
                )
                if sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    header_dict = {}
                    for row in sheet.iter_rows(min_row=start_r, max_row=end_r):
                        for col_idx, cell in enumerate(row, start=1):
                            val = (
                                str(cell.value).strip()
                                if cell.value is not None
                                else ""
                            )
                            if val:
                                header_dict.setdefault(col_idx, []).append(val)

                    for col_idx, parts in header_dict.items():
                        full_name = " - ".join(parts)
                        col_letter = get_column_letter(col_idx)
                        original_name = full_name
                        counter = 1
                        while full_name in self.available_headers:
                            full_name = f"{original_name} ({counter})"
                            counter += 1
                        self.available_headers[full_name] = col_letter
                        self.letter_to_name[col_letter] = full_name
            except (
                PermissionError,
                openpyxl.utils.exceptions.InvalidFileException,
                FileNotFoundError,
                OSError,
            ):
                pass
            finally:
                if wb is not None:
                    wb.close()

        self.headers_refreshed.emit(self.letter_to_name)

    def load_data(self, config_data: dict, current_profile: str, all_profiles: list):
        self.cmb_profile.blockSignals(True)
        self.cmb_profile.clear()
        self.cmb_profile.addItems(all_profiles)
        self.cmb_profile.setCurrentText(current_profile)
        self.cmb_profile.blockSignals(False)

        target_path = config_data.get("target_file", {}).get("path", "")
        self.txt_target_path.setText(target_path)

        saved_sheet = config_data.get("target_file", {}).get("sheet_name", "BrewSync")
        self._populate_sheet_names(target_path, saved_sheet)

        self.txt_header_row.setText(str(config_data.get("header_row", 5)))
        self.txt_input_sheet.setText(
            config_data.get("input_file", {}).get("sheet_name", "")
        )
        self.txt_fingerprint_rule.setText(config_data.get("fingerprint", ""))

        fallback_profiles = [""] + [p for p in all_profiles if p != current_profile]
        self.cmb_fallback_profile.clear()
        self.cmb_fallback_profile.addItems(fallback_profiles)
        self.cmb_fallback_profile.setCurrentText(
            config_data.get("fallback_profile", "")
        )

        self._refresh_headers()
        self._update_chain_preview()

    def get_data(self) -> dict:
        h_text = self.txt_header_row.text().strip()
        if not re.match(r"^\d+(-\d+)?$", h_text):
            raise ValueError("Dòng tiêu đề phải là số (vd: 5) hoặc khoảng (vd: 1-3)!")

        return {
            "profile_name": self.cmb_profile.currentText().strip(),
            "target_file": {
                "path": self.txt_target_path.text().strip(),
                "sheet_name": self.cmb_sheet_name.currentText().strip(),
            },
            "input_file": {"sheet_name": self.txt_input_sheet.text().strip()},
            "header_row": h_text,
            "fingerprint": self.txt_fingerprint_rule.text().strip(),
            "fallback_profile": self.cmb_fallback_profile.currentText(),
        }
