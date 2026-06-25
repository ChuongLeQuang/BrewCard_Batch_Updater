"""
EN: Excel sync service for updating Master file.
VI: Dịch vụ đồng bộ Excel để cập nhật file Tổng.
"""

import os
import shutil
from datetime import datetime
import openpyxl
from openpyxl.utils import column_index_from_string
from typing import List, Dict, Any, Optional
from src.models.config_model import AppConfig
from src.models.excel_data_model import BrewRecord
from src.config.constants import FORMAT_OPTIONS


class BrewSyncEngine:
    """EN: Engine handling Insert/Update logic. VI: Động cơ xử lý chèn/cập nhật."""

    def __init__(self, config: AppConfig):
        self.config = config
        self.target_path = self.config.data["target_file"]["path"]
        self.sheet_name = self.config.data["target_file"]["sheet_name"]
        self.mappings = self.config.data["mappings"]
        self.batch_col_letter = self._find_batch_column()
        self.last_backup_path = ""

    def _find_batch_column(self) -> str:
        """
        EN: Find the primary key column (Explicitly marked by user).
        VI: Lấy chữ cái của cột Khóa chính (Được người dùng tích chọn).
        """
        for m in self.mappings:
            if m.get("is_key"):
                return m.get("target_col_letter")
        return self.mappings[0].get("target_col_letter", "A") if self.mappings else "A"

    def _backup_master_file(self) -> str:
        """
        EN: Automatically backup the master file before overwriting.
        VI: Tự động sao lưu tệp đích trước khi thực hiện ghi đè.
        """
        if not os.path.exists(self.target_path):
            return ""

        backup_dir = "data/backups"
        os.makedirs(backup_dir, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        profile_name = (
            self.config.profile_name
            if hasattr(self.config, "profile_name")
            else "Profile"
        )
        backup_name = f"Master_Backup_{profile_name}_{timestamp}.xlsx"
        backup_path = os.path.join(backup_dir, backup_name)

        shutil.copy2(self.target_path, backup_path)
        self.last_backup_path = backup_path
        return backup_path

    def _find_row_by_batch(self, ws, col_idx: int, batch_num: str) -> Optional[int]:
        """
        EN: Find row index by batch number.
        VI: Dò tìm chỉ số dòng dựa trên số mẻ nấu (khóa chính).
        """
        header_row_val = str(self.config.data.get("header_row", "1"))
        if "-" in header_row_val:
            header_end_idx = int(header_row_val.split("-")[1])
        else:
            header_end_idx = int(header_row_val)
        data_start_row = header_end_idx + 1

        batch_str_target = str(batch_num).strip().lower()
        if not batch_str_target:
            return None

        # Duyệt từ dòng data_start_row đến max_row
        for row in range(data_start_row, ws.max_row + 1):
            val = ws.cell(row=row, column=col_idx).value
            if val is not None:
                if str(val).strip().lower() == batch_str_target:
                    return row
        return None

    def _find_next_empty_row(self, ws, col_idx: int) -> int:
        """
        EN: Find next empty row for inserting new record, checking for safety.
        VI: Tìm dòng trống tiếp theo để chèn mẻ nấu mới, có kiểm tra an toàn.
        """
        header_row_val = str(self.config.data.get("header_row", "1"))
        if "-" in header_row_val:
            header_end_idx = int(header_row_val.split("-")[1])
        else:
            header_end_idx = int(header_row_val)
        data_start_row = header_end_idx + 1

        row = data_start_row
        while True:
            val = ws.cell(row=row, column=col_idx).value
            if val is None or str(val).strip() == "":
                # Kiểm tra thêm 4 dòng tiếp theo để tránh dòng trống xen kẽ
                is_real_empty = True
                for offset in range(1, 5):
                    next_val = ws.cell(row=row + offset, column=col_idx).value
                    if next_val is not None and str(next_val).strip() != "":
                        is_real_empty = False
                        break
                if is_real_empty:
                    return row
            row += 1

    def sync_records(self, records: List[BrewRecord]) -> bool:
        """
        EN: Sync records into the target master file (Insert/Update in-place).
        VI: Đồng bộ danh sách dữ liệu vào file tổng (Thêm mới/Cập nhật tại chỗ).
        """
        # 1. Load or create workbook
        is_new_file = not os.path.exists(self.target_path)
        if not is_new_file:
            self._backup_master_file()
            wb = openpyxl.load_workbook(self.target_path)
            ws = (
                wb[self.sheet_name]
                if self.sheet_name in wb.sheetnames
                else wb.create_sheet(self.sheet_name)
            )
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = self.sheet_name

        # Phân tích dòng tiêu đề từ cấu hình (vd: "5" hoặc "1-3")
        header_row_val = str(self.config.data.get("header_row", "1"))
        if "-" in header_row_val:
            header_end_idx = int(header_row_val.split("-")[1])
        else:
            header_end_idx = int(header_row_val)

        batch_col_idx = column_index_from_string(self.batch_col_letter)
        col_letter_to_idx = {
            m["target_col_letter"]: column_index_from_string(m["target_col_letter"])
            for m in self.mappings
        }
        col_letter_to_format = {
            m["target_col_letter"]: FORMAT_OPTIONS.get(
                m.get("format_type", "📝 Mặc định"), "General"
            )
            for m in self.mappings
        }

        # Tạo tự động Header nếu là file hoàn toàn mới (hoặc sheet hoàn toàn trống)
        if is_new_file or (ws.max_row <= 1 and ws.cell(1, 1).value is None):
            for m in self.mappings:
                c_idx = column_index_from_string(m["target_col_letter"])
                ws.cell(row=header_end_idx, column=c_idx, value=m.get("target_col", ""))

        # 2. Merge records in-place
        for record in records:
            batch_num = str(record.batch_number).strip()

            row_idx = self._find_row_by_batch(ws, batch_col_idx, batch_num)
            if row_idx is None:
                row_idx = self._find_next_empty_row(ws, batch_col_idx)
                ws.cell(row=row_idx, column=batch_col_idx, value=batch_num)

            for letter, value in record.data.items():
                if letter in col_letter_to_idx:
                    c_idx = col_letter_to_idx[letter]
                    cell = ws.cell(row=row_idx, column=c_idx, value=value)
                    fmt = col_letter_to_format.get(letter, "General")
                    if fmt != "General":
                        cell.number_format = fmt

        # 3. Save workbook
        wb.save(self.target_path)
        wb.close()
        return True
